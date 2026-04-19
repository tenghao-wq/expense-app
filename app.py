import re
import io
from datetime import datetime
import streamlit as st
import pdfplumber
from openpyxl import load_workbook

st.set_page_config(page_title="发票报销单生成器", page_icon="🧾", layout="centered")
st.title("🧾 发票报销单生成器")
st.caption("上传发票 PDF，自动识别并填写报销单模板")

DATA_START_ROW = 11
DATA_END_ROW = 26

EXPENSE_TYPE_MAP = {
    "住宿": "住宿", "餐饮": "餐饮", "餐费": "餐饮", "饮食": "餐饮",
    "交通": "交通", "运输": "交通", "网约车": "交通", "出租车": "交通", "滴滴": "交通",
    "高铁": "差旅", "火车": "差旅", "机票": "差旅", "航空": "差旅",
    "办公": "办公用品", "会议": "会议", "培训": "培训", "广告": "广告", "服务": "服务费",
}

def classify_expense(text):
    for keyword, category in EXPENSE_TYPE_MAP.items():
        if keyword in text:
            return category
    return "其他"

def extract_invoice_info(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception:
        return None

    if not text.strip():
        return None

    invoice_type = "增值税专用发票" if "专用发票" in text else "普通发票"

    date_val = None
    m = re.search(r"开票日期[：:]\s*(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if m:
        try:
            date_val = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    amount = None
    m = re.search(r"[（(]小写[）)]\s*[¥￥]?\s*([\d,]+\.?\d*)", text)
    if m:
        try:
            amount = float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    if amount is None:
        m = re.search(r"价税合计.*?[¥￥]([\d,]+\.?\d*)", text)
        if m:
            try:
                amount = float(m.group(1).replace(",", ""))
            except ValueError:
                pass

    content = ""
    m = re.search(r"\*([\u4e00-\u9fa5A-Za-z]+)\*([\u4e00-\u9fa5A-Za-z]+)", text)
    if m:
        content = f"{m.group(1)}-{m.group(2)}"

    expense_type = classify_expense(content + text[:200])

    return {
        "date": date_val,
        "amount": amount,
        "invoice_type": invoice_type,
        "invoice_content": content,
        "expense_type": expense_type,
    }

def fill_template(template_bytes, invoices):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    ranges_to_unmerge = [
        str(r) for r in ws.merged_cells.ranges
        if r.min_row >= DATA_START_ROW and r.max_row <= DATA_END_ROW
    ]
    for r in ranges_to_unmerge:
        ws.unmerge_cells(r)

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{r}"].value = None

    invoices_sorted = sorted(invoices, key=lambda x: x["date"] or datetime.min)

    for i, inv in enumerate(invoices_sorted[:DATA_END_ROW - DATA_START_ROW + 1]):
        r = DATA_START_ROW + i
        ws[f"B{r}"] = inv["date"]
        ws[f"C{r}"] = inv["expense_type"]
        ws[f"D{r}"] = inv["amount"]
        ws[f"E{r}"] = inv["invoice_type"]
        ws[f"F{r}"] = inv["invoice_content"]
        ws[f"G{r}"] = 1
        ws[f"H{r}"] = ""

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ── UI ──────────────────────────────────────────────

col1, col2 = st.columns(2)

with col1:
    st.subheader("① 上传报销单模板")
    template_file = st.file_uploader("选择 .xlsx 模板文件", type=["xlsx"], key="template")

with col2:
    st.subheader("② 上传发票")
    invoice_files = st.file_uploader(
        "选择发票 PDF（可多选）", type=["pdf"], accept_multiple_files=True, key="invoices"
    )

if template_file and invoice_files:
    if st.button("生成报销单", type="primary", use_container_width=True):
        template_bytes = template_file.read()

        invoices = []
        failed = []
        progress = st.progress(0, text="识别中...")

        for i, f in enumerate(invoice_files):
            info = extract_invoice_info(f.read())
            if info:
                invoices.append(info)
            else:
                failed.append(f.name)
            progress.progress((i + 1) / len(invoice_files), text=f"识别中：{f.name}")

        progress.empty()

        if invoices:
            result = fill_template(template_bytes, invoices)
            total = sum(inv["amount"] for inv in invoices if isinstance(inv.get("amount"), (int, float)))

            st.success(f"识别成功 {len(invoices)} 张，合计 ¥{total:,.2f}")

            if failed:
                st.warning(f"以下文件无法识别（可能是扫描件）：{', '.join(failed)}")

            st.subheader("识别明细")
            rows = []
            for inv in sorted(invoices, key=lambda x: x["date"] or datetime.min):
                rows.append({
                    "日期": inv["date"].strftime("%Y-%m-%d") if inv["date"] else "—",
                    "类型": inv["expense_type"],
                    "金额": f"¥{inv['amount']:,.2f}" if inv["amount"] else "—",
                    "发票内容": inv["invoice_content"],
                    "发票类型": inv["invoice_type"],
                })
            st.dataframe(rows, use_container_width=True, hide_index=True)

            month = datetime.now().strftime("%Y-%m")
            st.download_button(
                label="⬇️ 下载报销单",
                data=result,
                file_name=f"报销单_{month}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        else:
            st.error("所有发票均无法识别，请确认是否为电子发票 PDF。")
else:
    st.info("请先上传模板和发票文件。")
